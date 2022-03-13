from exchangelib import DELEGATE, Configuration, Credentials, Account, FileAttachment
import os
import PyPDF2
import re
import access
import openpyxl
from datetime import datetime

# This script writes to TRACKER.xlsx

def dateConvert(d):
    return datetime.strptime(d, '%B %d, %Y').strftime('%m/%d/%Y')
def readEmail(name, emails):
    for item in emails:
        body = (str(item.text_body))
        time = re.search(r'(Sent:)(.*)\d\d\d\d', body).group()
        timeR = time[6:].split(' ', 1)[-1]
        timeSlice = dateConvert(timeR)
        subject = item.subject
        serviceOrder = int(re.search(r'\d{9}',subject).group())
        print(serviceOrder)
        wb = openpyxl.load_workbook('TRACKER.xlsx')
        sheet = wb['AMCOR']
        for rowNum in range(4, sheet.max_row):
            serviceOrders = sheet.cell(row=rowNum, column=5).value
            if serviceOrder == serviceOrders:
                sheet.cell(row=rowNum, column=6).value = name
        wb.save('TRACKER.xlsx')
credentials = Credentials(access.username, access.pwd)
config = Configuration(server='outlook.office365.com', credentials=credentials)
account = Account(primary_smtp_address='tylerosata@gmail.com', config=config, autodiscover=False, access_type=DELEGATE)
local_path = 'C:\\Users\\otyle\\spam\\'

folder1 = account.root/'Top of Information Store'/'Prepress'
prepress = folder1.all().order_by('-datetime_received')[:100]
folder2 = account.root/'Top of Information Store'/'Stepping'
stepping = folder2.all().order_by('-datetime_received')[:100]
folder3 = account.root/'Top of Information Store'/'Plating'
plating = folder3.all().order_by('-datetime_received')[:100]
folder4 = account.root/'Top of Information Store'/'Plates Shipped'
shipped = folder4.all().order_by('-datetime_received')[:100]
folder6 = account.root/'Top of Information Store'/'To Blue'
to_blue = folder6.all().order_by('-datetime_received')[:100]
folder7 = account.root/'Top of Information Store'/'Invoiced'
invoiced = folder7.all().order_by('-datetime_received')[:100]

readEmail('Prepress', prepress)
readEmail('Stepping', stepping)
readEmail('Plating', plating)

for item in shipped:
    for attachment in item.attachments:
        if 'Order Confirmation Form.pdf' in str(attachment.name):
            if isinstance(attachment, FileAttachment):
                download_path = os.path.join(local_path, attachment.name)
                with open(download_path, 'wb') as f:
                    f.write(attachment.content)
                    pdfFileObj = open('Order Confirmation Form.pdf', 'rb')
                    pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
                    pageObj = pdfReader.getPage(0)
                    page = pageObj.extractText()
                    order = re.compile(r'Service Order No.\d{9}')
                    date = re.compile(r'Delivery Date\d\d/\d\d/\d\d\d\d')
                    service = str(order.search(page))
                    ship = str(date.search(page))
                    delivery = re.compile(r'\d\d/\d\d/\d\d\d\d')
                    last = re.compile(r'\d{9}')
                    shipDate = delivery.search(ship).group()
                    serviceNumber = last.search(service).group()
                    number = int(serviceNumber)
                    pdfFileObj.close()
                    print(serviceNumber)
                    wb = openpyxl.load_workbook('TRACKER.xlsx')
                    sheet = wb['AMCOR']
                    for rowNum in range(4, sheet.max_row):
                        serviceOrders = sheet.cell(row=rowNum, column=5).value
                        if number == serviceOrders:
                            sheet.cell(row=rowNum, column=6).value = 'Shipped'
                    wb.save('TRACKER.xlsx')

for item in invoiced:
    for attachment in item.attachments:
        if '.pdf' in str(attachment.name):
            if isinstance(attachment, FileAttachment):
                download_path = os.path.join(local_path, 'invoice.pdf')
                with open(download_path, 'wb') as f:
                    f.write(attachment.content)
                    pdfFileObj = open('invoice.pdf', 'rb')
                    pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
                    pageObj = pdfReader.getPage(0)
                    page = pageObj.extractText()
                    order = re.search(r'Sales Order No.\d{9}', page).group()
                    salesOrder = int(re.search(r'\d{9}',order).group())
                    print(str(salesOrder) + ' is invoiced.')
                    pdfFileObj.close()
                    wb = openpyxl.load_workbook('TRACKER.xlsx')
                    sheet = wb['AMCOR']
                    for rowNum in range(2, sheet.max_row):
                        salesorders = sheet.cell(row=rowNum, column=4).value
                        if salesOrder == salesorders:
                            sheet.cell(row=rowNum, column=6).value = 'Invoiced'
                    wb.save('TRACKER.xlsx')
