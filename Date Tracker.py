from exchangelib import DELEGATE, Configuration, Credentials, Account, FileAttachment
import os
import PyPDF2
import re
import access
import openpyxl
from datetime import datetime

#This script writes to Tracker_test.xlsx


def dateConvert(d):
    return datetime.strptime(d, '%B %d, %Y').strftime('%m/%d/%Y')
def readEmail(columns, emails, folder):
    to_folder = folder/'archive'
    for item in emails:
        body = (str(item.text_body))
        time = re.search(r'(Sent:)(.*)\d\d\d\d', body).group()
        timeR = time[6:].split(' ', 1)[-1]
        timeSlice = dateConvert(timeR)
        subject = item.subject
        serviceOrder = int(re.search(r'\d{9}',subject).group())
        print(serviceOrder)
        wb = openpyxl.load_workbook('Tracker_test.xlsx')
        sheet = wb['tracker']
        for rowNum in range(5, sheet.max_row):
            serviceOrders = sheet.cell(row=rowNum, column=6).value
            if serviceOrder == serviceOrders:
                sheet.cell(row=rowNum, column=columns).value = timeSlice
        wb.save('Tracker_test.xlsx')
        item.is_read = True
        item.save()
        item.move(to_folder)
    
credentials = Credentials(access.username, access.pwd)
config = Configuration(server='outlook.office365.com', credentials=credentials)
account = Account(primary_smtp_address='tylerosata@gmail.com', config=config, autodiscover=False, access_type=DELEGATE)
local_path = 'C:\\Users\\otyle\\spam\\'

folder1 = account.root/'Top of Information Store'/'Prepress'
prepress = folder1.all().order_by('-datetime_received')[:100]
folder2 = account.root/'Top of Information Store'/'Stepping'
stepping = folder2.all().order_by('-datetime_received')[:100]
folder3 = account.root/'Top of Information Store'/'Plating'
plating = folder3.all().order_by('-datetime_received')[:100]
folder4 = account.root/'Top of Information Store'/'Plates Shipped'
shipped = folder4.all().order_by('-datetime_received')[:100]
folder5 = account.root/'Top of Information Store'/'Amcor Approval'
printer_approval = folder5.all().order_by('-datetime_received')[:100]
folder7 = account.root/'Top of Information Store'/'Invoiced'
invoiced = folder7.all().order_by('-datetime_received')[:100]
folder8 = account.root/'Top of Information Store'/'Order Submitted'
submitted = folder8.all().order_by('-datetime_received')[:100]

readEmail(9, prepress, folder1)
readEmail(12, stepping, folder2)

for item in printer_approval:
    to_folder = folder5/'archive'
    body = (str(item.text_body))
    date = re.compile(r'(Sent:)(.*)\d\d\d\d')
    time = str(date.search(body).group())
    timeR = time[6:].split(' ', 1)[-1]
    timeSlice = dateConvert(timeR)
    subject = item.subject
    searcher1 = re.compile(r'\d{6}')
    mt_number = int(searcher1.search(subject).group())
    print(mt_number)
    wb = openpyxl.load_workbook('Tracker_test.xlsx')
    sheet = wb['tracker']
    for rowNum in range(5, sheet.max_row):
        MT_Numbers = sheet.cell(row=rowNum, column=3).value
        if mt_number == MT_Numbers:
            sheet.cell(row=rowNum, column=14).value = timeSlice
    wb.save('Tracker_test.xlsx')
    item.is_read = True
    item.save()
    item.move(to_folder)

readEmail(15, submitted, folder8)
readEmail(16, plating, folder3)

for item in shipped:
    to_folder = folder4/'archive'
    for attachment in item.attachments:
        if 'Order Confirmation Form.pdf' in str(attachment.name):
            if isinstance(attachment, FileAttachment):
                download_path = os.path.join(local_path, attachment.name)
                with open(download_path, 'wb') as f:
                    f.write(attachment.content)
                    pdfFileObj = open('Order Confirmation Form.pdf', 'rb')
                    pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
                    pageObj = pdfReader.getPage(0)
                    page = pageObj.extractText()
                    order = re.compile(r'Service Order No.\d{9}')
                    date = re.compile(r'Delivery Date\d\d/\d\d/\d\d\d\d')
                    service = str(order.search(page))
                    ship = str(date.search(page))
                    delivery = re.compile(r'\d\d/\d\d/\d\d\d\d')
                    last = re.compile(r'\d{9}')
                    shipDate = delivery.search(ship).group()
                    serviceNumber = last.search(service).group()
                    number = int(serviceNumber)
                    pdfFileObj.close()
                    print(serviceNumber)
                    wb = openpyxl.load_workbook('Tracker_test.xlsx')
                    sheet = wb['tracker']
                    for rowNum in range(5, sheet.max_row):
                        serviceOrders = sheet.cell(row=rowNum, column=6).value
                        if number == serviceOrders:
                            sheet.cell(row=rowNum, column=17).value = shipDate
                    wb.save('Tracker_test.xlsx')
    item.is_read = True
    item.save()
    item.move(to_folder)

for item in invoiced:
    to_folder = folder7/'archive'
    for attachment in item.attachments:
        if '.pdf' in str(attachment.name):
            if isinstance(attachment, FileAttachment):
                download_path = os.path.join(local_path, 'invoice.pdf')
                with open(download_path, 'wb') as f:
                    f.write(attachment.content)
                    pdfFileObj = open('invoice.pdf', 'rb')
                    pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
                    pageObj = pdfReader.getPage(0)
                    page = pageObj.extractText()
                    date = re.search(r'\d\d/\d\d/\d\d\d\d',page).group()
                    print(date)
                    order = re.search(r'Sales Order No.\d{9}', page).group()
                    salesOrder = int(re.search(r'\d{9}',order).group())
                    print(salesOrder)
                    pdfFileObj.close()
                    wb = openpyxl.load_workbook('Tracker_test.xlsx')
                    sheet = wb['tracker']
                    for rowNum in range(2, sheet.max_row):
                        salesorders = sheet.cell(row=rowNum, column=5).value
                        if salesOrder == salesorders:
                            sheet.cell(row=rowNum, column=18).value = date
                    wb.save('Tracker_test.xlsx')
    item.is_read = True
    item.save()
    item.move(to_folder)
